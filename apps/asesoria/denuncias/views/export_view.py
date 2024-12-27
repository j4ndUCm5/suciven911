from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ..models import Denuncia


class DenunciaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = ""

    def get(self, request, *args, **kwargs):
        # Filtra los datos del modelo para generar el archivo Excel
        denuncias = Denuncia.objects.all()

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:P1")  # Mezcla las celdas de A1 a D1
        ws["A1"] = "Denuncias del 911"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega el texto del título

        ws.append(
            [
                "Estatus",
                "Ente",
                "Nombres del denunciante",
                "Apellidos del denunciante",
                "Cedula del denunciante",
                "Telefono",
                "Email",
                "Direccion",
                "Nombres del denunciado",
                "Apellidos del denunciado",
                "Cedula del denunciado",
                "Motivo",
                "Zona",
                "Fecha de la denuncia",
                "Fecha del incidente",
            ]
        )  # Reemplaza con los nombres de tus campos
        # Obtiene las columnas y establece el ancho deseado
        columnas = ws.column_dimensions
        columnas["A"].width = 10
        columnas["B"].width = 15
        columnas["C"].width = 20
        columnas["D"].width = 20
        columnas["E"].width = 25
        columnas["F"].width = 25
        columnas["G"].width = 25
        columnas["H"].width = 25
        columnas["I"].width = 25
        columnas["J"].width = 25
        columnas["K"].width = 25
        columnas["L"].width = 25
        columnas["M"].width = 25
        columnas["N"].width = 25
        columnas["O"].width = 20
        columnas["P"].width = 20

        for dato in denuncias:
            ws.append(
                [
                    dato.estatus,
                    dato.ente,
                    dato.nombres_d,
                    dato.apellidos_d,
                    dato.cedula_d,
                    dato.telefono,
                    dato.email,
                    dato.direccion_d,
                    dato.nombres_denunciado,
                    dato.apellidos_denunciado,
                    dato.cedula_denunciado,
                    dato.motivo,
                    dato.zona,
                    dato.fecha_denuncia,
                    dato.fecha_incidente,
                ]
            )  # Reemplaza con los campos de tu modelo

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Denuncias.xlsx")
