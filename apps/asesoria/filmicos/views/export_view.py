from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ..models import RegistroFilmico


class RegistroFilmicoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = ""

    def get(self, request, *args, **kwargs):
        # Filtra los datos del modelo para generar el archivo Excel
        registro = RegistroFilmico.objects.all()

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:G1")  # Mezcla las celdas de A1 a D1
        ws["A1"] = "Registros Filmicos del 911"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega el texto del título

        ws.append(
            [
                "Estatus",
                "Direccion",
                "Camara",
                "Motivo de solicitud",
                "Ente que solicita",
                "Fecha de la solicitud",
                "Fecha de culminacion",
            ]
        )  # Reemplaza con los nombres de tus campos
        # Obtiene las columnas y establece el ancho deseado
        columnas = ws.column_dimensions
        columnas["A"].width = 15
        columnas["B"].width = 20
        columnas["C"].width = 20
        columnas["D"].width = 20
        columnas["E"].width = 30
        columnas["F"].width = 30
        columnas["G"].width = 30

        for dato in registro:
            ws.append(
                [
                    dato.estatus,
                    dato.direccion,
                    dato.camara,
                    dato.motivo_solicitud,
                    dato.ente_solicita,
                    dato.fecha_solicitud,
                    dato.fecha_culminacion,
                ]
            )  # Reemplaza con los campos de tu modelo

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="RegistroFilmico.xlsx")
